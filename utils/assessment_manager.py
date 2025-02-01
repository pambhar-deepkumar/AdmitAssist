import json
from dataclasses import dataclass

import openpyxl
from openpyxl.utils.cell import range_boundaries

from .enums import TestResult


@dataclass
class RowData:
    """
    Represents a single row in an Excel sheet corresponding to a course/module.

    Attributes:
        name (str): The name of the course/module.
        credits (float): The number of credits associated with the course/module.
        evaluation (TestResult, optional): The evaluation result for the course/module.
        comment (str, optional): A comment regarding the evaluation.
        eval_cell (str, optional): The cell reference for the evaluation result in the Excel sheet.
        comment_cell (str, optional): The cell reference for the comment in the Excel sheet.

    Methods:
        set_evaluation(evaluation: TestResult, comment: str):
            Sets the evaluation result and comment for this row.
    """

    name: str
    credits: float
    evaluation: TestResult = None
    comment: str = None
    eval_cell: str = None
    comment_cell: str = None

    def set_evaluation(self, evaluation: TestResult, comment: str):
        """
        Sets the evaluation result and comment for this row.

        Args:
            evaluation (TestResult): The evaluation result.
            comment (str): A comment regarding the evaluation.
        """
        self.evaluation = evaluation
        self.comment = comment


class ModuleData:
    """
    Represents a module in an Excel sheet containing multiple rows of data.

    Attributes:
        name (str): The name of the module.
        content (str): A description or content summary of the module.
        learnOut (str): The learning outcomes of the module.
        maxCred (int): The maximum credits that can be earned from this module.
        rows (list[RowData]): A list of RowData objects representing rows in this module.

    Methods:
        __iter__():
            Initializes an iterator over the rows in the module.

        __next__():
            Returns the next row in the module during iteration.
    """

    def __init__(
        self, name: str, content: str, learnOut: str, maxCred: int, rows: list[RowData]
    ):
        """
        Initializes a ModuleData object.

        Args:
            name (str): The name of the module.
            content (str): A description or content summary of the module.
            learnOut (str): The learning outcomes of the module.
            maxCred (int): The maximum credits that can be earned from this module.
            rows (list[RowData]): A list of RowData objects representing rows in this module.
        """
        self.name = name
        self.content = content
        self.learnOut = learnOut
        self.maxCred = maxCred
        self.rows = rows

    def __iter__(self):
        """
        Initializes an iterator over the rows in the module.

        Returns:
            ModuleData: The current instance to enable iteration over its rows.
        """
        self._index = 0
        return self

    def __next__(self):
        """
        Returns the next row in the module during iteration.

        Returns:
            RowData: The next row in the module.

        Raises:
            StopIteration: If there are no more rows to iterate through.
        """
        while self._index < len(self.rows):
            row = self.rows[self._index]
            self._index += 1
            if row.name is not None:
                return row
        raise StopIteration


class AssessmentManager:
    """
    Manages assessment data for multiple modules in an Excel sheet.

    Attributes:
        data (dict): The assessment data loaded from a JSON file.
        wb (Workbook): The Excel workbook object.
        ws (Worksheet): The active worksheet object within the workbook.
        modules (list[ModuleData]): A list of ModuleData objects representing all modules.

    Methods:
        _process_modules():
            Processes modules from JSON data and creates ModuleData objects.

        _create_module_rows(module):
            Creates RowData objects for each row in a given module's ranges.

        _get_cell_references(cell_range):
            Generates a list of cell references from a given range string.

        save(path):
            Saves all evaluations and comments back to their respective locations in an Excel file.

        get_wb():
            Returns the current workbook object with all updates applied.

        __iter__():
            Initializes an iterator over all modules.

        __next__():
            Returns the next module during iteration.
    """

    def __init__(self, json_file, excel_file):
        """
        Initializes an AssessmentManager object by loading data from JSON and Excel files.

        Args:
            json_file (str): Path to the JSON file containing assessment data.
            excel_file (str): Path to the Excel file containing course/module information.

        Raises:
            FileNotFoundError: If either JSON or Excel file is not found at specified paths.
            ValueError: If JSON data is improperly formatted or missing required fields.
        """
        print("Initializing AssessmentManager...")
        with open(json_file, "r") as f:
            self.data = json.load(f)
            print("Loaded JSON data.")

        self.wb = openpyxl.load_workbook(excel_file)
        print(f"Loaded Excel workbook from {excel_file}.")
        self.ws = self.wb.active
        print(f"Active worksheet is '{self.ws.title}'.")
        self.modules = []
        self._process_modules()

    def _process_modules(self):
        """
        Processes the modules from the loaded JSON data and creates ModuleData objects.

        This method reads each section of the JSON data to extract information about modules,
        such as their name, content, learning outcomes, and maximum credits. For each module,
        it creates a list of RowData objects representing rows in the module and appends
        a new ModuleData object to the `modules` attribute.

        Raises:
            KeyError: If required keys are missing in the JSON data for any module.

        Example:
            A module in the JSON might look like this:
            {
                "Course": "Module Name",
                "Content": "Module Content",
                "Learning Outcome": "Learning Outcome Description",
                "Minimum Credits": 5,
                "Name Location": "B2:B10",
                "Credits Location": "C2:C10",
                "Evaluation Result Location": "D2:D10",
                "Comments Location": "E2:E10"
            }
        """
        for section in self.data.values():
            modules = section.get("Modules") or section.get("modules") or []
            for module in modules:
                module_name = module["Course"]
                module_content = module["Content"]
                module_learning_outcome = module["Learning Outcome"]
                module_max_credit = module["Minimum Credits"]
                rows = self._create_module_rows(module)
                self.modules.append(
                    ModuleData(
                        module_name,
                        module_content,
                        module_learning_outcome,
                        module_max_credit,
                        rows,
                    )
                )

    def _create_module_rows(self, module):
        """
        Creates RowData objects for each row in a given module's ranges.

        This method uses cell range information provided in the JSON data to extract
        values from specific cells in the Excel sheet. It generates RowData objects
        that represent individual rows within a module.

        Args:
            module (dict): A dictionary containing information about a single module,
                           including cell ranges for names, credits, evaluations, and comments.

                           Example keys include:
                           - "Name Location"
                           - "Credits Location"
                           - "Evaluation Result Location"
                           - "Comments Location"

        Returns:
            list[RowData]: A list of RowData objects representing rows in this module.

        Raises:
            KeyError: If required keys like "Name Location" or "Credits Location" are missing.

        Example:
            If the cell ranges are as follows:
                Name Location: B2:B5
                Credits Location: C2:C5
                Evaluation Result Location: D2:D5
                Comments Location: E2:E5

            This method will create 4 RowData objects with values extracted from these ranges.
        """
        name_cells = self._get_cell_references(module["Name Location"])
        credit_cells = self._get_cell_references(module["Credits Location"])
        eval_cells = self._get_cell_references(module["Evaluation Result Location"])
        comment_cells = self._get_cell_references(module["Comments Location"])

        return [
            RowData(
                name=self.ws[name].value,
                credits=self.ws[credit].value,
                eval_cell=eval,
                comment_cell=comment,
            )
            for name, credit, eval, comment in zip(
                name_cells, credit_cells, eval_cells, comment_cells
            )
        ]

    def _get_cell_references(self, cell_range):
        """
        Returns a list of cell references (e.g., ['B32', 'B33', ...]) from a given range.
        """
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        return [
            f"{openpyxl.utils.get_column_letter(min_col)}{row}"
            for row in range(min_row, max_row + 1)
        ]

    def save(self, path):
        """
        Saves all evaluations and comments back to their respective locations in the Excel file.

        This method iterates over all modules and their rows to update the corresponding
        cells in the Excel sheet with evaluation results and comments. It then saves the
        modified workbook to the specified file path.

        Args:
            path (str): The file path where the updated Excel workbook should be saved.

        Example:
            manager.save("updated_assessment.xlsx")

        Raises:
            PermissionError: If the file cannot be written to (e.g., due to insufficient permissions).
        """
        print("Saving updates to Excel...")
        for module in self.modules:
            for row in module:
                if row.eval_cell and row.evaluation:
                    self.ws[row.eval_cell] = row.evaluation.value
                if row.comment_cell and row.comment:
                    self.ws[row.comment_cell] = row.comment
        self.wb.save(path)

    def get_wb(self):
        """
        Returns the current workbook object with all updates applied.

        This method ensures that all evaluations and comments are written to their respective
        cells in the workbook before returning it.

        Returns:
            Workbook: The updated workbook object.

        Example:
            wb = manager.get_wb()
            wb.save("final_assessment.xlsx")
        """
        for module in self.modules:
            for row in module:
                if row.eval_cell and row.evaluation:
                    self.ws[row.eval_cell] = row.evaluation.value
                if row.comment_cell and row.comment:
                    self.ws[row.comment_cell] = row.comment
        return self.wb

    def __iter__(self):
        """
        Initializes an iterator over all modules.

        This allows the `AssessmentManager` object to be used in a `for` loop
        to iterate over its modules.

        Returns:
            AssessmentManager: The current instance, ready for iteration.

        Example:
            for module in manager:
                print(module.name)
                for row in module:
                    print(row.name, row.credits)
        """
        self._index = 0
        return self

    def __next__(self):
        if self._index < len(self.modules):
            module = self.modules[self._index]
            self._index += 1
            return module
        raise StopIteration
